import numpy as np
from openpyxl import load_workbook
from sklearn.metrics.pairwise import cosine_similarity
from typing import Set, Iterable, List
from spacy.tokens import Span
import langdetect
import regex as re
import os

COLOR_DICT = {
    (0, 0, 0): "nvt",
    (255, 146, 208): "yes",
    (255, 255, 0): "no",
    (255, 255, 235): "maybe",
    (255, 255, 199): "no",
    (255, 198, 239): "yes",
}


def overwrite_spans(spans: Iterable["Span"]) -> List["Span"]:
    """Filter a sequence of spans and remove duplicates or overlaps. Useful for
    creating named entities (where one token can only be part of one entity) or
    when merging spans with `Retokenizer.merge`. When spans overlap, the (first)
    longest span is preferred over shorter spans.

    spans (Iterable[Span]): The spans to filter.
    RETURNS (List[Span]): The filtered spans.
    """
    # get_sort_key = lambda span: (span.end - span.start, -span.start)
    # sorted_spans = sorted(spans, key=get_sort_key, reverse=True)
    result = []
    seen_tokens: Set[int] = set()
    for span in spans:
        # Check for end - 1 here because boundaries are inclusive
        if span.start not in seen_tokens and span.end - 1 not in seen_tokens:
            result.append(span)
            seen_tokens.update(range(span.start, span.end))
    result = sorted(result, key=lambda span: span.start)
    return result


def color_tranlate_helper(input_file):
    """
    Helper function to translate green, yellow and red to yes/no. Uses rbg dict as input and finds closest match
    using cosine similarity. Hopefully to account for slight variations in color.

    Parameters
    ----------
    input_file : str
        Input excel file

    Returns
    -------
    val: list
        List of dictionaries containing {'vraag': vraagnr, 'kleur': rgb, 'antwoord': answer}
    """

    if isinstance(input_file, str):

        wb = load_workbook(input_file, data_only=True)
        sh = wb[wb.sheetnames[0]]
        column_names = np.asarray([col[0].value for col in sh.columns])
        col_idx = np.where(column_names == "Contract dVb")[0][0]
        column = list(sh.columns)[col_idx]

        val = []
        for cell in column:
            vraagnr = sh["A" + str(cell.row)].value
            color_in_hex = (
                cell.fill.start_color.index
            )  # this gives you Hexadecimal value of the color
            try:
                rgb = tuple(int(color_in_hex[i : i + 2], 16) for i in (0, 2, 4))
                check_rgb = rgb  # Color in RGB
            except TypeError:
                rgb = (0, 0, 0)
                check_rgb = rgb
            if check_rgb == (0, 0, 0):
                if str(cell.value).lower() in ["ja", "yes", "true"]:
                    check_rgb = (255, 146, 208)
                elif str(cell.value).lower() in ["nee", "geen", "niet gevonden"]:
                    check_rgb = (255, 255, 0)
                elif str(cell.value).lower() not in ["", "nan", "none"]:
                    check_rgb = (255, 146, 208)
                else:
                    check_rgb = (0, 0, 0)
            answer = [v for v in COLOR_DICT.values()][
                np.argmax(
                    cosine_similarity([check_rgb], [k for k in COLOR_DICT.keys()])
                )
            ]
            val.append(
                {
                    "vraag": vraagnr,
                    "kleur": rgb,
                    "check_rgb": check_rgb,
                    "antwoord": answer,
                }
            )

        return val


def load_files(DF, contract_nr, main_path):
    files = []
    for i, row in DF.loc[DF["contractnr"] == contract_nr].iterrows():
        try:
            path = os.path.join(
                main_path, str(row["output_path"]), str(row["ext_id"]) + ".txt"
            )
            with open(path, "r") as file:
                text = file.read().replace("\n", " ")
            files.append(text)
        except Exception as ex:
            print("File not found: ", ex)
    return files


def safe_langdetect(text):
    try:
        return langdetect.detect(text)
    except Exception as e:
        return ""


def clean_text(text):
    text = text.replace("{", "")
    text = text.replace(")", "")
    text = text.replace("(", "")
    text = text.replace("\n", " ")
    return text


def apply_regex(doc, pattern, label):

    new_ents = []

    for match in re.finditer(pattern, doc.text.lower()):
        start, end = match.span()
        span = doc.char_span(start, end)
        if span is not None:
            new_ent = Span(doc, span.start, span.end, label=label)
            new_ents.append(new_ent)

    original_ents = list(doc.ents)
    new_ents.extend(original_ents)

    doc.ents = overwrite_spans(new_ents)
    return doc
